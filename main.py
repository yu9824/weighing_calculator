'''
Copyright (c) 2021, yu9824
This software is released under the GNU LESSER GENERAL PUBLIC LICENSE Version 3 License, see LICENSE.
'''

from typing import Tuple
from element_recognition import get_ratio, make_compositions, element_recognition
from math import isclose
import pandas as pd
import numpy as np
import PySimpleGUI as sg
import os
import json
from glob import glob
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

'''
Mac OSX10.15においてmenubarがつかえないbugが起きるが，これはstandalone化すれば解決する．
'''

# フォント
font_default = 'Hiragino Maru Gothic ProN'
font_size_default = 19
option_text_default = {
    'font': (font_default, font_size_default),
}

APP_NAME = 'Weighing Calculator'

# 設定
path_root = os.path.abspath(os.path.dirname(__file__))
path_settings = os.path.join(path_root, 'settings.json')
path_atomic_weights = os.path.join(path_root, 'atomic_weights.csv')

class WeighingCalculator:
    def __init__(self, materials):
        '''
        Parameters
        ----------
        materials : list
            [description]
        '''
        self.materials = materials
        self.df_atomic_weights = pd.read_csv(path_atomic_weights, index_col = 0, comment='#')
        self.dict_materials = {material: self._get_formula_weight(material) for material in materials}

    def calc(self, products = [], ratio = [], mg = 2000, excess = {}, exact = True, progress_bar = True):
        '''
        calculate weighing

        Parameters
        ----------
        products : list, optional
            [description], by default []
        ratio : list, optional
            [description], by default []
        mg : int, optional
            完成量, by default 2000
        excess : dict, optional
            過剰量 e.g. {'Li2O': 0.05}, by default {}
        exact : bool, optional
            完全一致していないとダメかどうか, by default True
        progress_bar : bool, optional
            進捗バーを表示するかどうか, by default True

        Raises
        ------
        ValueError
        '''

        self.mg = mg

        if len(products) * len(ratio):  # 両方に入力があったら．
            raise ValueError('You can only enter either "products" or "ratio".')
        elif len(products):
            self.df_ratio = get_ratio(materials=self.materials, products = products, exact = exact)
        elif len(ratio):
            products = make_compositions(self.materials, ratio = ratio).index.to_numpy().tolist()
            if isinstance(ratio, pd.DataFrame):
                ratio = ratio.loc[:, self.materials].to_numpy()
            elif isinstance(ratio, (pd.Series, np.ndarray, list)):
                ratio = np.array(ratio).reshape(-1, len(self.materials))
            self.df_ratio = pd.DataFrame(ratio, columns = self.materials, index = products)
        else:
            raise ValueError('You have to enter either "products" or "ratio".')

        # 比率が計算できなかった組成をためておくリスト
        # comp_null = []
        
        products = self.df_ratio.index.to_numpy().tolist()  # self.df_ratioではproductsの空白を削除した組成名を得られるため，上書き．
        self.dict_products = {product: self._get_formula_weight(product) for product in products}
        self.products = products

        self.excess = excess

        # 結果を足すSeries
        srs_material_weight = []
        srs_material_weight_excess = []

        # moleをとっておくリスト
        self.moles = []

        # dataframeを各行ごとに取り出す．
        for product, ratio in self.df_ratio.iterrows():
            mole = mg / self.dict_products[product]
            self.moles.append(mole)
            if ratio.isnull().all():    # anyのほうが良い気もする．
                pass
            else:   # きちんと計算できていたら
                sr_material_weight = pd.Series(ratio * mole * np.array([self.dict_materials[material] for material in self.materials]), index = self.materials, name = product)

                # 過剰量の計算
                sr_material_weight_excess = sr_material_weight.copy()
                for material, excess_ratio in self.excess.items():
                    sr_material_weight_excess[material] *= (1 + excess_ratio)

                srs_material_weight.append(sr_material_weight.to_frame().transpose())
                srs_material_weight_excess.append(sr_material_weight_excess.to_frame().transpose())
        self.df_material_weight = pd.concat(srs_material_weight, axis = 0, sort = False)
        self.df_material_weight_excess = pd.concat(srs_material_weight_excess, axis = 0, sort = False)
    
    def _get_formula_weight(self, formula:str) -> float:
        """get formula weight

        Parameters
        ----------
        formula : str
            chemical formula, e.g.) 'Li2O'

        Returns
        -------
        float
            formula weight
        """
        return (element_recognition(formula, elements = self.df_atomic_weights.index.tolist()).values @ self.df_atomic_weights.values)[0][0]
        


class gui:
    def __init__(self, theme = 'LightGray1'):
        '''
        theme: 'LightGray1' or 'Black'
        '''
        # load setting
        try:
            self.settings = json.load(open(path_settings, mode = 'r', encoding = 'utf_8'))
        except Exception as e:
            pass
        else:
            if self.settings['theme'] == 'light':
                theme = 'LightGray1'
            elif self.settings['theme'] == 'dark':
                theme = 'Black'
        finally:
            sg.theme(theme)

        self.lang = self.settings['lang']

        # constants
        self.threshold_scroll = 4
        self.fname_cache_materials = 'cache_materials.json'
        self.fname_lang = 'lang.json'

        if os.path.isfile(self.fname_lang):
            with open(self.fname_lang, mode = 'r', encoding = 'utf_8') as f:
                self.lang_dict = json.load(f)
         

    def run(self):
        # オブジェクトの生成
        menu_n_materials = Menu()

        # layoutの作成
        menu_n_materials.layout = [
            [sg.Text(self.lang_dict[self.lang]['start_menu'], font = (font_default, 25))],
            [sg.Text('')],
            [sg.Text(self.lang_dict[self.lang]['n_materials']), sg.InputText(key = 'n_materials', size = (3, 1), justification='right')],
            [sg.Text('')],
            [sg.Submit(self.lang_dict[self.lang]['next'], key = 'Next')]
        ]

        # windowの生成
        menu_n_materials.make_window()
        while True:
            menu_n_materials.read()
            if menu_n_materials.event is None:
                # windowを閉じて抜ける
                menu_n_materials.window.close()
                break
            elif menu_n_materials.event == 'Next':
                n_materials = menu_n_materials.values['n_materials']
                try:
                    n_materials = int(n_materials)
                except ValueError:
                    popup_error = sg.popup_error('You have to enter an integer.', **option_text_default, modal = False, keep_on_top=True)
                else:
                    # windowを閉じてから次のwindowを開く
                    menu_n_materials.window.close()
                    self._get_materials_compositions(n_materials)
                    break
    
    def _load_cache_materials(self, n_materials:int)->list:
        """load_cache_materials

        Parameters
        ----------
        n_materials : int
            how many materials

        Returns
        -------
        list
            cached materials or empty list (length = n_materials)
        """
        if os.path.isfile(self.fname_cache_materials):
            with open(self.fname_cache_materials, mode = 'r', encoding = 'utf_8') as f:
                chache_materials:dict = json.load(f)
            if str(n_materials) in chache_materials:    # converted to str because of json format.
                return chache_materials[str(n_materials)]
        return ['' for _ in range(n_materials)]
        
    def _dump_cache_materials(self, materials:list):
        """dump_cache_materials

        Parameters
        ----------
        materials : list
            materials

        Returns
        -------
        None
        """
        n_materials = len(materials)
        if os.path.isfile(self.fname_cache_materials):
            with open(self.fname_cache_materials, mode = 'r', encoding = 'utf_8') as f:
                chache_materials:dict = json.load(f)
        else:
            chache_materials = {}
        chache_materials[str(n_materials)] = materials
        with open(self.fname_cache_materials, mode = 'w', encoding = 'utf_8') as f:
            json.dump(chache_materials, f, indent=4)
        
                

    def _get_materials_compositions(self, n_materials):
        # オブジェクトの生成
        get_materials_compositions = Menu()

        # 原料を入力するやつ
        entering_compositions_layout = [[sg.Text('{0}{1}'.format(self.lang_dict[self.lang]['material'], n+1), size = (8, 1)), sg.InputText(default_text=cached_material, size = (10, 1), key = 'material{}'.format(n+1))] for n, cached_material in enumerate(self._load_cache_materials(n_materials))]

        # このページのlayoutの作成
        if n_materials > self.threshold_scroll: # 数が多いときはスクロールできるように．
            get_materials_compositions.layout = [
                [sg.Column(entering_compositions_layout, scrollable = True, vertical_scroll_only=True, size = (400, 225), justification='center')],
            ]
        else:
            get_materials_compositions.layout = entering_compositions_layout
        # buttonとタイトルは共通なので足す．
        get_materials_compositions.layout.insert(0, [sg.Text(self.lang_dict[self.lang]['starting_materials_input'])])
        get_materials_compositions.layout.extend([[sg.Text('')], [sg.Submit(self.lang_dict[self.lang]['confirm'], key='Confirm')]])

        # windowの生成
        get_materials_compositions.make_window()
        while True:
            # read
            get_materials_compositions.read()
            if get_materials_compositions.event is None:
                # windowを閉じて抜ける
                get_materials_compositions.window.close()
                break
            elif get_materials_compositions.event == 'Confirm':
                dict_materials = {k:v for k, v in get_materials_compositions.values.items() if 'material' in str(k)}
                if '' in dict_materials.values():
                    sg.popup_error('Some of them are not filled in at all.', **option_text_default, modal = False, keep_on_top=True)
                    continue
                # windowを閉じてから次のwindowを開く
                get_materials_compositions.window.close()
                self._dump_cache_materials(materials=list(dict_materials.values()))
                self._calculation_menu(dict_materials)
                break


    def _calculation_menu(self, dict_materials):
        calculation_menu = Menu()

        n_materials = len(dict_materials)

        # これから定義する二つのフレームを同じように整えるoptions
        options_frame = {
            'border_width':0,
            'element_justification':'center',
            # 'size' : (300, 300)
        }

        # 原料比を入力するやつ
        layout_entering_ratio = [[sg.Text(dict_materials['material{}'.format(n+1)], size = (6, 1)), sg.InputText(size = (6, 1), key = dict_materials['material{}'.format(n+1)], justification='right')] for n in range(n_materials)]

        # ratio
        options_layout_ratio = {
            'size': (220, 170),
            'justification': 'center'
        }

        if n_materials > self.threshold_scroll: # スクロールできるように．
            layout_ratio = [
                [sg.Column(layout_entering_ratio, scrollable = True, vertical_scroll_only=True, **options_layout_ratio)],
            ]
        else:
            layout_ratio = layout_entering_ratio
        # buttonは共通なので足す．
        layout_ratio.extend([[sg.Text('')], [sg.Submit(self.lang_dict[self.lang]['calc_ratio'], key = 'Calc_ratio')]])

        frame_ratio = sg.Frame('', layout = layout_ratio, **options_frame)

        # 生成物を入力するやつ
        layout_product = [
            [sg.Text(self.lang_dict[self.lang]['calc_from_product'])],
            [sg.InputText(size = (15, 1), key = 'product')],
            [sg.Text('')],
            [sg.Submit(self.lang_dict[self.lang]['calc_product'], key = 'Calc_product')],
        ]
        frame_product = sg.Frame('', layout = layout_product,  **{k:v for k, v in options_frame.items() if k != 'border_widht'})

        # 過剰量を入力するやつ
        layout_excess = [[sg.Text(dict_materials['material{}'.format(n+1)], size = (6, 1)), sg.InputText('0.00', key = '{}_excess'.format(dict_materials['material{}'.format(n+1)]), size = (7, 1), justification='right'), sg.Text('mol%')] for n in range(n_materials)]
        if n_materials > self.threshold_scroll: # スクロールできるように
            layout_excess = [[sg.Column(layout_excess, scrollable=True, vertical_scroll_only=True, justification='center')]]
        # layout_excess.insert(0, [sg.Text('過剰量', justification='center')])
            
        frame_excess = sg.Frame(self.lang_dict[self.lang]['excess'], layout = layout_excess, **{k:v for k, v in options_frame.items() if k != 'border_width'})

        # 完成量を入力するやつ
        layout_mg = [
            [sg.Text(self.lang_dict[self.lang]['theoretical_amount'])],
            [sg.InputText('2000', key = 'mg', size = (7, 1), justification='right'), sg.Text('mg')]
        ]
        frame_mg = sg.Frame('', layout=layout_mg, **options_frame)

        # このページ全体のlayout
        calculation_menu.layout = [
            [sg.Text(self.lang_dict[self.lang]['calc_screen'])],
            [frame_mg],
            [sg.Text('')],
            [frame_ratio, sg.Text(' '), frame_product],
            [sg.Text('_'*100)],
            [sg.Text(self.lang_dict[self.lang]['common_settings'], justification='left')],
            [frame_excess]
        ]
        
        calculation_menu.make_window(size = (1200, 775), resizable = True)
        while True:
            calculation_menu.read()
            if calculation_menu.event is None:
                break
            if 'Calc' in calculation_menu.event:
                # インスタンス生成
                wc = WeighingCalculator(materials=list(dict_materials.values()))

                # 過剰量を辞書まとめる．
                dict_excess = {material: float(calculation_menu.values['{}_excess'.format(material)]) / 100 for material in dict_materials.values()}
                
                # 生成重量を変数として得る．
                mg = float(calculation_menu.values['mg'])
                if 'ratio' in calculation_menu.event:
                    try:
                        dict_ratio = {}
                        for k, v in calculation_menu.values.items():
                            if k in dict_materials.values():
                                if v.count('/') > 1:
                                    raise ValueError
                                elif '/' in v:
                                    dict_ratio[k] = float(v.split('/')[0]) / float(v.split('/')[1])
                                else:
                                    dict_ratio[k] = float(v)
                    except ValueError:  # try内で指定したValueError以外も含めて．
                        sg.popup_error('You have not entered any. Or the value you entered is not good.\nCorrect: 1/3, 1, 1.0, 3.141 etc.', **option_text_default, modal = False, keep_on_top=True)
                        continue
                    wc.calc(ratio = list(dict_ratio.values()), mg = mg, excess = dict_excess, exact = True, progress_bar = False)
                elif 'product' in calculation_menu.event:
                    if calculation_menu.values['product'] == '':
                        sg.popup_error('Nothing has been entered.', **option_text_default, modal = False, keep_on_top=True)
                        continue
                    try:    # exact=Trueでうまくいかなかったとき
                        wc.calc(products=[calculation_menu.values['product']], mg = mg, excess = dict_excess, exact = True, progress_bar = False)
                    except ValueError:
                        try:
                            wc.calc(products=[calculation_menu.values['product']], mg = mg, excess = dict_excess, exact = False, progress_bar = False)
                        except ValueError:  # exact=Falseでもうまくいかなかったとき
                            sg.PopupError('The composition you have entered is invalid.', **option_text_default, modal = False, keep_on_top=True)
                            continue
                        else:
                            sg.popup_ok('The results of the calculations may be different because they did not match exactly.', **option_text_default, modal = False, keep_on_top=True)

                self._table(wc = wc)
        
        calculation_menu.window.close()

    def _table(self, wc):
        '''
        wc: WeighingCalculatorオブジェクト
        '''
        # 出力用の表を作成
        df_output, df_output_show = self._make_output(wc)
        df_output_show.drop(index = ['measured value (mg)'], inplace = True)
        df_output.index.name = ''
        df_output_show.index.name = ''
        df_output_show.reset_index(inplace=True)

        # オブジェクトの生成
        table_menu = Menu()

        # layoutの作成
        table_menu.layout = [
            [sg.Table(df_output_show.to_numpy().tolist(), headings = df_output_show.columns.to_numpy().tolist(), col_widths = [16] + [10] * (df_output_show.shape[1]-2) + [16], auto_size_columns = False, hide_vertical_scroll=False)],
            [sg.Cancel(self.lang_dict[self.lang]['cancel']), sg.InputText(key='SaveAs', do_not_clear=False, enable_events=True, visible=False), sg.FileSaveAs(self.lang_dict[self.lang]['save_as'], file_types = (('Excel file', '*.xlsx'),))],   # do_not_clearをFalseにしないとCancelしても上書きされてしまう．あと，sg.FilseSaveAsのenable_event = Trueにしてもenentが発生しないのでsg.InputTextで受け取るときにeventを発生させている．
        ]

        # windowの作成
        table_menu.make_window(size = (1200, 600))

        while True:
            table_menu.read()
            if table_menu.event in (None, 'Cancel'):
                pass
            elif table_menu.event == 'SaveAs':
                if table_menu.values['SaveAs'] == '':
                    continue
                if os.path.splitext(os.path.basename(table_menu.values['SaveAs']))[-1] != '.xlsx':
                    fpath_output = os.path.splitext(table_menu.values['SaveAs'])[0] + '.xlsx'
                else:
                    fpath_output = table_menu.values['SaveAs']
                df_output.reset_index().to_excel(fpath_output, index = False)
                sg.PopupOK('Saved successfully.', modal = True, **option_text_default)
            break
            
                
        table_menu.window.close()

    def _make_output(self, wc: WeighingCalculator) -> Tuple[pd.DataFrame, pd.DataFrame]:
        '''
        wc: WeighingCalculatorオブジェクト
        '''
        cols = wc.materials + wc.products
        ind_molar_weight = 'M.W.'
        ind_molar_ratio = 'molar ratio'
        ind_mole = 'mole (mmol)'
        ind_excess_ratio = 'excess ratio (mol%)'
        ind_mole_with_excess = 'mole w/ excess'
        ind_weight_without_excess = 'no excess weight (mg)'
        ind_weight = 'weight (mg)'
        ind = [ind_molar_weight, ind_molar_ratio, ind_mole, ind_excess_ratio, ind_mole_with_excess, ind_weight_without_excess, ind_weight, 'measured value (mg)']
        
        df_output = pd.DataFrame([[np.nan] * len(cols)] * len(ind), columns = cols, index = ind)

        # 式量を代入
        for material in wc.materials:
            df_output.loc[ind_molar_weight, material] = wc.dict_materials[material]
        df_output.loc[ind_molar_weight, wc.products[0]] = wc.dict_products[wc.products[0]]

        # モル比を代入
        df_output.loc[ind_molar_ratio, wc.df_ratio.columns] = wc.df_ratio.values
        df_output.loc[ind_molar_ratio, wc.products] = 1

        # 表示用のDataFrameを作成
        df_output_show = df_output.copy()

        # モル数を代入
        # 以下使用する`+2`はそれぞれindex,columnsが占める+1と，pythonの0-indexをexcelの1-indexに変換する+1の合計．
        df_output_show.loc[ind_mole] = df_output.loc[ind_molar_ratio] * wc.moles[0]
        df_output.loc[ind_mole, wc.products] = '={0}{1}/{0}{2}'.format(
            get_column_letter(df_output_show.columns.get_loc(wc.products[0])+2),
            df_output.index.get_loc(ind_weight_without_excess)+2,
            df_output.index.get_loc(ind_molar_weight)+2,
        )
        df_output.loc[ind_mole, wc.materials] = [
            '={0}{2}/${1}${2}*${1}${3}'.format(
                get_column_letter(df_output.columns.get_loc(material)+2),
                get_column_letter(df_output.columns.get_loc(wc.products[0])+2),
                df_output.index.get_loc(ind_molar_ratio)+2,
                df_output.index.get_loc(ind_mole)+2,
            ) for material in wc.materials
        ]

        # 過剰量
        for material, ex in wc.excess.items():
            df_output.loc[ind_excess_ratio, material] = ex * 100
            df_output_show.loc[ind_excess_ratio, material] = ex * 100

        # 過剰量込のモル数
        df_output_show.loc[ind_mole_with_excess] = df_output_show.loc[ind_mole] * (1 + df_output_show.loc[ind_excess_ratio] / 100)
        df_output.loc[ind_mole_with_excess, wc.materials] = [
            '={0}{1}*(1+{0}{2}/100)'.format(
                get_column_letter(df_output.columns.get_loc(material)+2),
                df_output.index.get_loc(ind_mole)+2,
                df_output.index.get_loc(ind_excess_ratio)+2
            ) for material in wc.materials
        ]
        # ↓これ必要？
        for material in wc.materials:
            if np.isnan(df_output_show.loc[ind_mole_with_excess, material]):
                df_output_show.loc[ind_mole_with_excess, material] = df_output_show.loc[ind_mole, material]

        # 過剰量なしの重量
        df_output_show.loc[ind_weight_without_excess, wc.df_material_weight.columns] = wc.df_material_weight.loc[wc.products[0]]
        df_output.loc[ind_weight_without_excess, wc.materials] = [
            '={0}{1}*{0}{2}'.format(
                get_column_letter(df_output.columns.get_loc(material)+2),
                df_output.index.get_loc(ind_molar_weight)+2,
                df_output.index.get_loc(ind_mole)+2,
            ) for material in wc.materials
        ]

        df_output_show.loc[ind_weight_without_excess, wc.products[0]] = wc.mg
        df_output.loc[ind_weight_without_excess, wc.products[0]] = wc.mg

        # 過剰量ありの重量
        df_output_show.loc[ind_weight, wc.df_material_weight_excess.columns] = wc.df_material_weight_excess.loc[wc.products[0]]
        df_output.loc[ind_weight, wc.materials] = [
            '={0}{1}*{0}{2}'.format(
                get_column_letter(df_output.columns.get_loc(material)+2),
                df_output.index.get_loc(ind_molar_weight)+2,
                df_output.index.get_loc(ind_mole_with_excess)+2,
            ) for material in wc.materials
        ]

        df_output_show.loc[ind_weight, wc.products[0]] = df_output_show.loc[ind_weight, wc.df_material_weight_excess.columns].sum()
        df_output.loc[ind_weight, wc.products[0]] = '=SUM({0}{2}:{1}{2})'.format(
            get_column_letter(df_output.columns.get_loc(wc.materials[0])+2),
            get_column_letter(df_output.columns.get_loc(wc.materials[-1])+2),
            df_output.index.get_loc(ind_weight)+2,
        )

        # 表示用のdf_outputを作成
        df_output_show = df_output_show.applymap(lambda x:'{:.2f}'.format(x))
        df_output_show.replace('nan', '', inplace=True)

        return df_output, df_output_show




class Menu:
    def __init__(self, layout = ((sg.Text('You have to add something to show.')),)):
        self.layout = layout
        self.menu_def = [
            ['Menu', ['About {}'.format(APP_NAME), '---', 'Setting', 'Clear cache', '---', 'Exit']],
        ]

    def make_window(self, **options):
        default_options = {
            'size' : (800, 450),
            'element_justification' : 'center',
            'resizable': True
        }
        default_options.update(option_text_default)
        default_options.update(**options)
        # Add MenuBar
        self.layout.insert(0, [sg.Menu(self.menu_def, font = sg.DEFAULT_FONT)])
        self.window = sg.Window(APP_NAME, layout = self.layout, **default_options, finalize = True)

    def read(self):
        self.event, self.values = self.window.read()
        if self.event == 'Setting':
            if _change_setting():   # 設定に反映させるために閉じる場合
                self.window.close()
        elif self.event == 'Clear cache':
            _clear_cache()
        elif self.event == 'Exit':
            if sg.PopupOKCancel('Are you sure you want to exit?', **option_text_default) == 'OK':
                exit()
        elif self.event == 'About {}'.format(APP_NAME):
            with open('about.txt', mode = 'r', encoding = 'utf_8') as f:
                lcns = f.read()
            sg.PopupOK(lcns, modal= False, keep_on_top= True, title = 'About {}'.format(APP_NAME))
            
            
    
def _change_setting():
    settings = json.load(open(path_settings, mode = 'r', encoding = 'utf_8'))

    corr_lang = {
        '日本語': 'ja',
        'English': 'en',
    }
    corr_lang.update({v:k for k, v in corr_lang.items()})

    corr_theme = {
        'Light': 'light',
        'Dark': 'dark',
    }
    corr_theme.update({v:k for k, v in corr_theme.items()})

    setting_menu = Menu(layout = [
        [sg.Text('Setting')],
        [sg.Text('')],
        [sg.Text('Langage'), sg.Combo(['日本語', 'English'], default_value = corr_lang[settings['lang']], key = 'lang')],
        [sg.Text('Theme'), sg.Combo(['Light', 'Dark'], default_value = corr_theme[settings['theme']], key = 'theme')],
        [sg.Text('')],
        [sg.Cancel(), sg.OK()]
    ])

    setting_menu.make_window(size = (None, None))
    while True:
        setting_menu.read()
        do_close = False    # 再起動するかどうか
        if setting_menu.event == 'OK':
            settings = {
                'lang': corr_lang[setting_menu.values['lang']],
                'theme': corr_theme[setting_menu.values['theme']]
                }
            event = sg.PopupYesNo('You will need to reboot to apply the configuration changes.\nCan I close it to apply the settings?', modal = False, keep_on_top = True, **option_text_default)
            if event == 'Yes':
                json.dump(settings, open(path_settings, mode = 'w', encoding='utf_8'), indent = 4)
                do_close = True
                break
        else:
            break
    setting_menu.window.close()
    return do_close


def _clear_cache():
    cache_files = glob(os.path.join(path_root, 'cache*.json'))
    if cache_files:
        if sg.PopupOKCancel('Do you want to clear cache?', modal = False, keep_on_top = True, **option_text_default) == 'OK':
            for cache_file in cache_files:
                os.remove(cache_file)
            else:
                sg.PopupOK('Cache cleared.', modal = False, keep_on_top = True, **option_text_default)
    else:
        sg.PopupOK('No cache.', modal = False, keep_on_top = True, **option_text_default)

            
if __name__ == '__main__':
    # from pdb import set_trace
    
    app = gui()
    app.run()

    # debug for output
    # materials = ['Li2O', 'SiO2', 'MoO3']
    # wc = WeighingCalculator(materials)
    # wc.calc(ratio=[1,1,1], excess={
    #     material: 0.
    #     for material in materials
    # })
    # app._make_output(wc)